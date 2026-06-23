#!/usr/bin/env python3
"""
Generate a deterministic real-world WTA scenario from waypoint_timeline.csv.

The scenario clock is a slice of the global attack timeline:
    local t = 0      -> global 01:30:00
    local t = 150 s  -> global 01:32:30

The output follows the same core JSON schema as scenario_001.json:
    weapon_infos
    target_infos
    probability_table
    assignment_request
    engagement_windows

Extra metadata fields are included for reproducibility and piecewise trajectory
visualization. Unknown top-level fields are ignored by the existing C++ loader.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


EPS = 1e-9
KM2M = 1000.0
MPS_PER_KMH = 1.0 / 3.6

WEAPON_CODE_MAP = {
    "AK-176": "AK176",
    "AK-630": "AK630",
    "TCDT BV15": "TCDT_BV15",
    "VCM-T": "VCM_T",
}

WEAPON_TYPE_MAP = {
    "AK-176": 2,
    "AK-630": 2,
    "TCDT BV15": 3,
    "VCM-T": 4,
}

TARGET_CODE_MAP = {
    "Ship": "TGT_SHIP",
    "YJ-83": "TGT_YJ83",
    "USV TN-0301": "TGT_USV",
    "USV TN-0302": "TGT_USV",
    "UAV TN-0201": "TGT_UAV_FIXED",
}

TARGET_INFO_DEFS = [
    {
        "ID": 1,
        "Code": "TGT_SHIP",
        "Description": "Enemy destroyer",
        "Type": 4,
    },
    {
        "ID": 2,
        "Code": "TGT_YJ83",
        "Description": "YJ-83 anti-ship missile",
        "Type": 1,
    },
    {
        "ID": 3,
        "Code": "TGT_USV",
        "Description": "Unmanned surface vehicle",
        "Type": 5,
    },
    {
        "ID": 4,
        "Code": "TGT_UAV_FIXED",
        "Description": "Fixed-wing UAV",
        "Type": 3,
    },
]

ENTITY_PRIORITY = {
    "Ship": 0,
    "YJ-83": 1,
    "USV TN-0301": 2,
    "USV TN-0302": 3,
    "UAV TN-0201": 4,
}


def parse_hms(value: str) -> float:
    parts = [float(part) for part in value.strip().split(":")]
    if len(parts) == 3:
        hours, minutes, seconds = parts
    elif len(parts) == 2:
        hours = 0.0
        minutes, seconds = parts
    elif len(parts) == 1:
        return parts[0]
    else:
        raise ValueError(f"Invalid time value: {value!r}")
    return hours * 3600.0 + minutes * 60.0 + seconds


def format_hms(seconds: float) -> str:
    total_ms = int(round(seconds * 1000.0))
    total_s, ms = divmod(total_ms, 1000)
    hours, rem = divmod(total_s, 3600)
    minutes, secs = divmod(rem, 60)
    if ms:
        return f"{hours:d}:{minutes:02d}:{secs:02d}.{ms:03d}"
    return f"{hours:d}:{minutes:02d}:{secs:02d}"


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value))
    return "".join(ch for ch in text if not unicodedata.combining(ch)).lower()


def canonical_target_name(entity: str, role: str) -> Optional[str]:
    norm = normalize_text(entity)
    role_norm = normalize_text(role)

    if role_norm == "enemy_ship" or "khu truc" in norm or "destroyer" in norm:
        return "Ship"
    if "yj-83" in norm or "yj83" in norm:
        return "YJ-83"
    if "tn-0301" in norm:
        return "USV TN-0301"
    if "tn-0302" in norm:
        return "USV TN-0302"
    if "tn-0201" in norm:
        return "UAV TN-0201"
    return None


def parse_numeric_range(
    value: object,
    *,
    full_default: Tuple[float, float],
) -> Tuple[float, float]:
    if value is None:
        return full_default

    text = str(value).strip().lower()
    if text == "full":
        return full_default

    match = re.search(
        r"(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)",
        text,
    )
    if not match:
        raise ValueError(f"Cannot parse range value: {value!r}")

    return float(match.group(1)), float(match.group(2))


def read_catalog(workbook_path: Path) -> dict:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["Sheet1"]

    weapon_names = [
        str(worksheet.cell(1, column).value).strip()
        for column in range(2, 6)
    ]
    weapon_column = {
        name: index + 2
        for index, name in enumerate(weapon_names)
    }

    row_by_label = {}
    for row in range(1, worksheet.max_row + 1):
        label = worksheet.cell(row, 1).value
        if label is not None:
            row_by_label[str(label).strip()] = row

    target_rows = {}
    for row in range(2, 8):
        name = worksheet.cell(row, 1).value
        if name:
            target_rows[str(name).strip()] = row

    max_shot_rows = {}
    # The pair-specific shot-cap table occupies rows 21-26.  The updated
    # workbook contains another route table below it with repeated entity names,
    # so do not scan to worksheet.max_row or those later rows will overwrite the
    # actual cap rows.
    for row in range(21, min(26, worksheet.max_row) + 1):
        name = worksheet.cell(row, 1).value
        if name:
            max_shot_rows[str(name).strip()] = row

    weapon_infos = []
    weapons = []
    probability_table = []
    pair_max_by_name = {}

    for weapon_id, weapon_name in enumerate(weapon_names, start=1):
        column = weapon_column[weapon_name]

        elevation_min, elevation_max = parse_numeric_range(
            worksheet.cell(row_by_label["Elevation"], column).value,
            full_default=(-90.0, 90.0),
        )
        azimuth_from, azimuth_to = parse_numeric_range(
            worksheet.cell(row_by_label["Azimuth"], column).value,
            full_default=(0.0, 360.0),
        )
        altitude_min_km, altitude_max_km = parse_numeric_range(
            worksheet.cell(row_by_label["Altitude"], column).value,
            full_default=(-1.0, 100.0),
        )

        azimuth_from %= 360.0
        if abs(azimuth_to - 360.0) > EPS:
            azimuth_to %= 360.0

        max_shots_values = []
        for target_name, row in max_shot_rows.items():
            value = worksheet.cell(row, column).value
            if isinstance(value, (int, float)):
                pair_max_by_name[(weapon_name, target_name)] = int(value)
                if value > 0:
                    max_shots_values.append(int(value))

        weapon_info = {
            "ID": weapon_id,
            "Code": WEAPON_CODE_MAP[weapon_name],
            "Type": WEAPON_TYPE_MAP[weapon_name],
            "MinRange": int(
                round(
                    float(
                        worksheet.cell(
                            row_by_label["MinRange"],
                            column,
                        ).value
                    )
                    * KM2M
                )
            ),
            "MaxRange": int(
                round(
                    float(
                        worksheet.cell(
                            row_by_label["MaxRange"],
                            column,
                        ).value
                    )
                    * KM2M
                )
            ),
            "MinAltitude": int(round(altitude_min_km * KM2M)),
            "MaxAltitude": int(round(altitude_max_km * KM2M)),
            "AzimuthFromDeg": azimuth_from,
            "AzimuthToDeg": azimuth_to,
            "ElevationMinDeg": elevation_min,
            "ElevationMaxDeg": elevation_max,
            # The original schema has only one cap per weapon type.
            # Preserve the largest pair-specific cap here and export the exact
            # pair caps separately in pair_max_shots.
            "MaxShotsPerTarget": max(max_shots_values, default=0),
            "RoundsPerBurst": int(
                worksheet.cell(
                    row_by_label["BurstInfo"],
                    column,
                ).value
            ),
            "BurstInterval": float(
                worksheet.cell(
                    row_by_label["BurstInterval"],
                    column,
                ).value
            ),
            "ReloadTime": float(
                worksheet.cell(
                    row_by_label["ReloadTime"],
                    column,
                ).value
            ),
        }
        weapon_infos.append(weapon_info)

        weapons.append(
            {
                "ID": weapon_id,
                "WTAVesselID": 1,
                "Ammo": int(
                    worksheet.cell(
                        row_by_label["Ammo"],
                        column,
                    ).value
                ),
                "WTAWeaponInfoCode": weapon_info["Code"],
                "Status": 1,
            }
        )

    threat_by_name = {}
    speed_by_name = {}

    for target_name, row in target_rows.items():
        threat_value = worksheet.cell(row, 7).value
        speed_value = worksheet.cell(row, 8).value
        if isinstance(threat_value, (int, float)):
            threat_by_name[target_name] = float(threat_value)
        if isinstance(speed_value, (int, float)):
            speed_by_name[target_name] = float(speed_value)

        target_code = TARGET_CODE_MAP.get(target_name)
        if target_code is None:
            continue

        for weapon_name in weapon_names:
            score = worksheet.cell(
                row,
                weapon_column[weapon_name],
            ).value
            if not isinstance(score, (int, float)) or score <= 0.0:
                continue
            probability_table.append(
                {
                    "Score": float(score),
                    "WTAWeaponInfoCode": WEAPON_CODE_MAP[weapon_name],
                    "WTATargetInfoCode": target_code,
                }
            )

    return {
        "weapon_infos": weapon_infos,
        "weapons": weapons,
        "probability_table": probability_table,
        "threat_by_name": threat_by_name,
        "catalog_speed_by_name": speed_by_name,
        "pair_max_by_name": pair_max_by_name,
        "weapon_name_by_code": {
            WEAPON_CODE_MAP[name]: name
            for name in weapon_names
        },
    }


def read_timeline(csv_path: Path) -> Dict[str, dict]:
    grouped: Dict[str, dict] = {}

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        required = {
            "entity",
            "role",
            "speed_kmh",
            "waypoint",
            "x_east_km",
            "y_north_km",
            "altitude_m",
            "global_time_s",
        }
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(
                f"Timeline CSV is missing columns: {sorted(missing)}"
            )

        for row in reader:
            entity = row["entity"].strip()
            role = row["role"].strip()
            canonical = canonical_target_name(entity, role)
            if canonical is None:
                continue

            record = grouped.setdefault(
                canonical,
                {
                    "entity": entity,
                    "role": role,
                    "canonical_name": canonical,
                    "speed_kmh": float(row["speed_kmh"]),
                    "launch_time_s": float(row.get("launch_time_s") or 0.0),
                    "launch_rule": row.get("launch_rule", ""),
                    "points": [],
                },
            )

            record["points"].append(
                {
                    "waypoint": row["waypoint"].strip(),
                    "global_time_s": float(row["global_time_s"]),
                    "x": float(row["x_east_km"]),
                    "y": float(row["y_north_km"]),
                    "z": float(row["altitude_m"]) / 1000.0,
                }
            )

    for record in grouped.values():
        record["points"].sort(
            key=lambda point: point["global_time_s"]
        )

    return grouped


def interpolate_point(
    points: Sequence[dict],
    global_time_s: float,
) -> dict:
    if not points:
        raise ValueError("Cannot interpolate an empty route.")

    if global_time_s <= points[0]["global_time_s"] + EPS:
        return dict(points[0])

    if global_time_s >= points[-1]["global_time_s"] - EPS:
        return dict(points[-1])

    for left, right in zip(points, points[1:]):
        t0 = left["global_time_s"]
        t1 = right["global_time_s"]
        if t0 - EPS <= global_time_s <= t1 + EPS:
            ratio = (global_time_s - t0) / (t1 - t0)
            return {
                "waypoint": f"{left['waypoint']}->{right['waypoint']}",
                "global_time_s": global_time_s,
                "x": left["x"] + ratio * (right["x"] - left["x"]),
                "y": left["y"] + ratio * (right["y"] - left["y"]),
                "z": left["z"] + ratio * (right["z"] - left["z"]),
            }

    raise RuntimeError("No interpolation segment found.")


def unit_direction(
    left: dict,
    right: dict,
) -> Tuple[float, float, float]:
    dx = right["X"] - left["X"]
    dy = right["Y"] - left["Y"]
    dz = right["Z"] - left["Z"]
    norm = math.sqrt(dx * dx + dy * dy + dz * dz)
    if norm <= EPS:
        return 0.0, 0.0, 0.0
    return dx / norm, dy / norm, dz / norm


def route_slice(
    record: dict,
    start_global_s: float,
    end_global_s: float,
) -> List[dict]:
    points = record["points"]
    route_start = max(start_global_s, points[0]["global_time_s"])
    route_end = min(end_global_s, points[-1]["global_time_s"])

    if route_end <= route_start + EPS:
        return []

    sliced = [interpolate_point(points, route_start)]

    for point in points:
        t = point["global_time_s"]
        if route_start + EPS < t < route_end - EPS:
            sliced.append(dict(point))

    sliced.append(interpolate_point(points, route_end))

    output = []
    for point in sliced:
        output.append(
            {
                "Time": round(
                    point["global_time_s"] - start_global_s,
                    6,
                ),
                "GlobalTime": round(point["global_time_s"], 6),
                "Waypoint": point["waypoint"],
                "X": round(point["x"], 6),
                "Y": round(point["y"], 6),
                "Z": round(point["z"], 6),
            }
        )

    # Remove accidental duplicate points at exact waypoint boundaries.
    deduplicated = []
    for point in output:
        if (
            deduplicated
            and abs(point["Time"] - deduplicated[-1]["Time"]) <= EPS
        ):
            deduplicated[-1] = point
        else:
            deduplicated.append(point)

    return deduplicated


def bearing_deg(x: float, y: float) -> float:
    return math.degrees(math.atan2(x, y)) % 360.0


def heading_vector(bearing: float) -> Tuple[float, float]:
    radians = math.radians(bearing)
    return math.sin(radians), math.cos(radians)


def angle_in_sector(
    angle: float,
    start: float,
    end: float,
) -> bool:
    span = (end - start) % 360.0
    if abs(span) <= 1e-8:
        return True
    relative = (angle - start) % 360.0
    return relative <= span + 1e-8


def quadratic_roots(
    a: float,
    b: float,
    c: float,
    duration: float,
) -> List[float]:
    roots = []

    if abs(a) > EPS:
        discriminant = b * b - 4.0 * a * c
        if discriminant >= -EPS:
            discriminant = max(0.0, discriminant)
            square_root = math.sqrt(discriminant)
            for root in (
                (-b - square_root) / (2.0 * a),
                (-b + square_root) / (2.0 * a),
            ):
                if -EPS <= root <= duration + EPS:
                    roots.append(min(max(root, 0.0), duration))
    elif abs(b) > EPS:
        root = -c / b
        if -EPS <= root <= duration + EPS:
            roots.append(min(max(root, 0.0), duration))

    return roots


def segment_intervals(
    start_point: dict,
    end_point: dict,
    weapon_info: dict,
    own_heading_deg: float,
) -> List[Tuple[float, float]]:
    t0 = float(start_point["Time"])
    t1 = float(end_point["Time"])
    duration = t1 - t0
    if duration <= EPS:
        return []

    x0 = float(start_point["X"])
    y0 = float(start_point["Y"])
    z0 = float(start_point["Z"])
    vx = (float(end_point["X"]) - x0) / duration
    vy = (float(end_point["Y"]) - y0) / duration
    vz = (float(end_point["Z"]) - z0) / duration

    range_min = float(weapon_info["MinRange"]) / KM2M
    range_max = float(weapon_info["MaxRange"]) / KM2M
    altitude_min = float(weapon_info["MinAltitude"]) / KM2M
    altitude_max = float(weapon_info["MaxAltitude"]) / KM2M
    azimuth_from = float(weapon_info["AzimuthFromDeg"])
    azimuth_to = float(weapon_info["AzimuthToDeg"])
    elevation_min = float(weapon_info["ElevationMinDeg"])
    elevation_max = float(weapon_info["ElevationMaxDeg"])

    candidates = [0.0, duration]

    velocity_squared = vx * vx + vy * vy + vz * vz
    dot_position_velocity = x0 * vx + y0 * vy + z0 * vz
    position_squared = x0 * x0 + y0 * y0 + z0 * z0

    for radius in (range_min, range_max):
        candidates.extend(
            quadratic_roots(
                velocity_squared,
                2.0 * dot_position_velocity,
                position_squared - radius * radius,
                duration,
            )
        )

    azimuth_span = (azimuth_to - azimuth_from) % 360.0
    full_azimuth = abs(azimuth_span) <= 1e-8

    if not full_azimuth:
        for relative_boundary in (azimuth_from, azimuth_to):
            absolute_boundary = (
                own_heading_deg + relative_boundary
            ) % 360.0
            angle = math.radians(absolute_boundary)
            denominator = vx * math.cos(angle) - vy * math.sin(angle)
            if abs(denominator) > EPS:
                root = -(
                    x0 * math.cos(angle)
                    - y0 * math.sin(angle)
                ) / denominator
                if -EPS <= root <= duration + EPS:
                    candidates.append(
                        min(max(root, 0.0), duration)
                    )

    full_elevation = (
        elevation_min <= -89.999
        and elevation_max >= 89.999
    )
    if not full_elevation:
        horizontal_velocity_squared = vx * vx + vy * vy
        horizontal_position_velocity = x0 * vx + y0 * vy
        horizontal_position_squared = x0 * x0 + y0 * y0

        for elevation in (elevation_min, elevation_max):
            tangent = math.tan(math.radians(elevation))
            tangent_squared = tangent * tangent
            candidates.extend(
                quadratic_roots(
                    vz * vz
                    - tangent_squared * horizontal_velocity_squared,
                    2.0
                    * (
                        z0 * vz
                        - tangent_squared
                        * horizontal_position_velocity
                    ),
                    z0 * z0
                    - tangent_squared * horizontal_position_squared,
                    duration,
                )
            )

    if abs(vz) > EPS:
        for altitude in (altitude_min, altitude_max):
            root = (altitude - z0) / vz
            if -EPS <= root <= duration + EPS:
                candidates.append(min(max(root, 0.0), duration))

    candidates = sorted(
        {
            round(min(max(candidate, 0.0), duration), 12)
            for candidate in candidates
        }
    )

    def in_envelope(local_time: float) -> bool:
        x = x0 + vx * local_time
        y = y0 + vy * local_time
        z = z0 + vz * local_time

        slant_range = math.sqrt(x * x + y * y + z * z)
        horizontal_range = math.hypot(x, y)
        if horizontal_range <= EPS:
            return False

        absolute_azimuth = bearing_deg(x, y)
        relative_azimuth = (
            absolute_azimuth - own_heading_deg
        ) % 360.0
        elevation = math.degrees(
            math.atan2(z, horizontal_range)
        )

        return (
            range_min - EPS
            <= slant_range
            <= range_max + EPS
            and altitude_min - EPS
            <= z
            <= altitude_max + EPS
            and angle_in_sector(
                relative_azimuth,
                azimuth_from,
                azimuth_to,
            )
            and elevation_min - EPS
            <= elevation
            <= elevation_max + EPS
        )

    intervals = []
    current_start: Optional[float] = None
    current_end: Optional[float] = None

    for left, right in zip(candidates, candidates[1:]):
        if right - left <= EPS:
            continue

        midpoint = (left + right) / 2.0

        if in_envelope(midpoint):
            if current_start is None:
                current_start = left
            current_end = right
        elif current_start is not None and current_end is not None:
            intervals.append(
                (t0 + current_start, t0 + current_end)
            )
            current_start = None
            current_end = None

    if current_start is not None and current_end is not None:
        intervals.append((t0 + current_start, t0 + current_end))

    return intervals


def merge_intervals(
    intervals: Iterable[Tuple[float, float]],
    tolerance: float = 1e-6,
) -> List[Tuple[float, float]]:
    sorted_intervals = sorted(intervals)
    merged: List[List[float]] = []

    for start, end in sorted_intervals:
        if not merged or start > merged[-1][1] + tolerance:
            merged.append([start, end])
        else:
            merged[-1][1] = max(merged[-1][1], end)

    return [(start, end) for start, end in merged]


def compute_target_intervals(
    trajectory: Sequence[dict],
    weapon_info: dict,
    own_heading_deg: float,
) -> List[Tuple[float, float]]:
    intervals = []

    for start_point, end_point in zip(
        trajectory,
        trajectory[1:],
    ):
        intervals.extend(
            segment_intervals(
                start_point,
                end_point,
                weapon_info,
                own_heading_deg,
            )
        )

    min_width = float(weapon_info["BurstInterval"])

    return [
        (start, end)
        for start, end in merge_intervals(intervals)
        if end - start + EPS >= min_width
    ]


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Generate a real WTA JSON scenario from a waypoint timeline CSV."
        )
    )
    parser.add_argument(
        "timeline",
        nargs="?",
        default="waypoint_output/waypoint_timeline.csv",
        help="Waypoint timeline CSV generated by plot_waypoint_timeline.py",
    )
    parser.add_argument(
        "--catalog",
        default="scenario.xlsx",
        help="Workbook containing weapon, probability, threat, and ammo data",
    )
    parser.add_argument(
        "--out",
        default="real_scenario_013000_013230.json",
        help="Output JSON path",
    )
    parser.add_argument(
        "--start",
        default="1:30:00",
        help="Global scenario start time, H:MM:SS",
    )
    parser.add_argument(
        "--end",
        default="1:32:30",
        help="Global scenario end time, H:MM:SS",
    )
    parser.add_argument(
        "--own-heading-deg",
        default="auto",
        help=(
            "Own-ship bearing clockwise from North, or 'auto' to face "
            "the enemy destroyer at scenario start"
        ),
    )
    parser.add_argument(
        "--defense-radius-km",
        type=float,
        default=25.0,
    )
    args = parser.parse_args()

    timeline_path = Path(args.timeline)
    catalog_path = Path(args.catalog)
    output_path = Path(args.out)

    start_global_s = parse_hms(args.start)
    end_global_s = parse_hms(args.end)
    if end_global_s <= start_global_s:
        raise ValueError("--end must be later than --start")

    horizon_s = end_global_s - start_global_s

    catalog = read_catalog(catalog_path)
    timeline = read_timeline(timeline_path)

    required_targets = [
        "Ship",
        "YJ-83",
        "USV TN-0301",
        "USV TN-0302",
        "UAV TN-0201",
    ]
    missing_targets = [
        name for name in required_targets if name not in timeline
    ]
    if missing_targets:
        raise ValueError(
            f"Timeline is missing required entities: {missing_targets}"
        )

    trajectories = {}
    for name in required_targets:
        trajectory = route_slice(
            timeline[name],
            start_global_s,
            end_global_s,
        )
        if len(trajectory) < 2:
            raise ValueError(
                f"{name} is not active for enough time in the selected horizon."
            )
        trajectories[name] = trajectory

    enemy_initial = trajectories["Ship"][0]

    if str(args.own_heading_deg).lower() == "auto":
        own_heading_deg = bearing_deg(
            enemy_initial["X"],
            enemy_initial["Y"],
        )
    else:
        own_heading_deg = float(args.own_heading_deg) % 360.0

    heading_x, heading_y = heading_vector(own_heading_deg)

    vessel = {
        "ID": 1,
        "X": 0.0,
        "Y": 0.0,
        "Z": 0.0,
        "Speed": 0.0,
        "HeadingX": round(heading_x, 6),
        "HeadingY": round(heading_y, 6),
        "HeadingZ": 0.0,
        "DefenseRadius": float(args.defense_radius_km),
    }

    target_names = sorted(
        required_targets,
        key=lambda name: ENTITY_PRIORITY[name],
    )

    targets = []
    target_id_by_name = {}

    for target_id, name in enumerate(target_names, start=1):
        target_id_by_name[name] = target_id
        record = timeline[name]
        trajectory = trajectories[name]

        first = trajectory[0]
        second = trajectory[1]
        vx, vy, vz = unit_direction(first, second)

        target = {
            "ID": target_id,
            "WTATargetInfoCode": TARGET_CODE_MAP[name],
            "X": first["X"],
            "Y": first["Y"],
            "Z": first["Z"],
            "VX": round(vx, 6),
            "VY": round(vy, 6),
            "VZ": round(vz, 6),
            # Timeline speeds are in km/h; the WTA JSON uses m/s.
            "Speed": round(record["speed_kmh"] * MPS_PER_KMH, 6),
            "ThreatScore": float(
                catalog["threat_by_name"][name]
            ),
        }
        targets.append(target)

    probability_lookup = {
        (
            row["WTAWeaponInfoCode"],
            row["WTATargetInfoCode"],
        ): row["Score"]
        for row in catalog["probability_table"]
    }

    weapon_info_by_code = {
        row["Code"]: row
        for row in catalog["weapon_infos"]
    }

    engagement_windows = {}
    all_engagement_intervals = {}
    pair_max_shots = {}

    for weapon in catalog["weapons"]:
        weapon_id = weapon["ID"]
        weapon_code = weapon["WTAWeaponInfoCode"]
        weapon_info = weapon_info_by_code[weapon_code]
        workbook_weapon_name = catalog["weapon_name_by_code"][
            weapon_code
        ]

        for target_name in target_names:
            target_id = target_id_by_name[target_name]
            target_code = TARGET_CODE_MAP[target_name]

            probability = probability_lookup.get(
                (weapon_code, target_code),
                0.0,
            )
            if probability <= 0.0:
                continue

            intervals = compute_target_intervals(
                trajectories[target_name],
                weapon_info,
                own_heading_deg,
            )
            if not intervals:
                continue

            key = f"{weapon_id}_{target_id}"

            # The core schema supports one interval. In this synchronized
            # scenario the intervals are normally contiguous; if not, use the
            # longest valid interval and retain all intervals in metadata.
            selected = max(
                intervals,
                key=lambda interval: (
                    interval[1] - interval[0],
                    -interval[0],
                ),
            )

            engagement_windows[key] = [
                round(selected[0], 6),
                round(selected[1], 6),
            ]
            all_engagement_intervals[key] = [
                [round(start, 6), round(end, 6)]
                for start, end in intervals
            ]

            exact_pair_cap = catalog["pair_max_by_name"].get(
                (workbook_weapon_name, target_name),
                0,
            )
            pair_max_shots[key] = int(exact_pair_cap)

    output = {
        "scenario_metadata": {
            "source_timeline": timeline_path.name,
            "source_catalog": catalog_path.name,
            "global_start_time": format_hms(start_global_s),
            "global_end_time": format_hms(end_global_s),
            "global_start_s": start_global_s,
            "global_end_s": end_global_s,
            "scenario_horizon_s": horizon_s,
            "own_ship_heading_deg": round(own_heading_deg, 6),
            "time_mapping": (
                "JSON target positions are evaluated at global 01:30:00; "
                "JSON engagement windows use local seconds in [0, 150]."
            ),
        },
        "weapon_infos": catalog["weapon_infos"],
        "target_infos": TARGET_INFO_DEFS,
        "probability_table": catalog["probability_table"],
        "assignment_request": {
            "vessels": [vessel],
            "weapons": catalog["weapons"],
            "targets": targets,
        },
        "engagement_windows": engagement_windows,
        # Non-breaking metadata used to preserve information that the original
        # core schema cannot represent directly.
        "pair_max_shots": pair_max_shots,
        "all_engagement_intervals": all_engagement_intervals,
        "target_trajectories": {
            str(target_id_by_name[name]): trajectories[name]
            for name in target_names
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        json.dump(output, file, indent=2, ensure_ascii=False)
        file.write("\n")

    print(f"Created: {output_path}")
    print(
        f"Global horizon: {format_hms(start_global_s)} -> "
        f"{format_hms(end_global_s)} ({horizon_s:g} s)"
    )
    print(
        f"Own-ship heading: {own_heading_deg:.3f} deg "
        f"(clockwise from North)"
    )
    print(f"Targets: {len(targets)}")
    print(f"Weapons: {len(catalog['weapons'])}")
    print(f"Engagement windows: {len(engagement_windows)}")

    for target in targets:
        print(
            f"  T{target['ID']} {target['WTATargetInfoCode']}: "
            f"pos=({target['X']:.3f}, {target['Y']:.3f}, {target['Z']:.3f}) km, "
            f"speed={target['Speed']:.3f} m/s"
        )


if __name__ == "__main__":
    main()

# python gen_data_real_scenario.py --start 1:30:00 --end 1:32:30 --own-heading-deg auto --out real_scenario_013000_013230.json