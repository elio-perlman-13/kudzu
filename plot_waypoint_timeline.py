#!/usr/bin/env python3
"""
Plot the moving enemy ship, launch events, target waypoint routes, and global
arrival times from the updated scenario workbook.

Operational schedule
--------------------
T0 = 00:00
    Enemy destroyer begins moving from D0.
    USV TN-0302 launches from the destroyer's D0 position.

T0 + 87 s
    USV TN-0301 launches from the destroyer's current moving position.

When the enemy destroyer reaches D1
    UAV TN-0201 launches from the destroyer's D1 position.

When the enemy destroyer reaches D2
    YJ-83 launches from the destroyer's D2 position.

Outputs
-------
- waypoint_timeline.csv
- timing_summary.csv
- waypoint_routes.png
- distance_to_own_ship.png
- launch_arrival_timeline.png

Coordinate convention
---------------------
Longitude W is negative, latitude N is positive. Local plot coordinates use
X = east displacement and Y = north displacement from the own ship, in km.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import matplotlib.pyplot as plt
from openpyxl import load_workbook


EARTH_RADIUS_KM = 6371.0088
WAYPOINT_LABELS = ("D0", "D1", "D2", "D3", "D4")

# User-provided approximate total travel data. These values are retained only
# for comparison; route geometry in the workbook is the source used to compute
# the plotted times.
ESTIMATED_TRAVEL = {
    "USV TN-0302": {"distance_km": 130.28, "travel_time_s": 100 * 60 + 13},
    "USV TN-0301": {"distance_km": 128.39, "travel_time_s": 98 * 60 + 46},
    "UAV TN-0201": {"distance_km": 122.8, "travel_time_s": 49 * 60 + 7},
    "Tên lửa YJ-83": {"distance_km": 70.6, "travel_time_s": 3 * 60 + 55},
}

# Launch rules. The enemy-waypoint rules are resolved after the destroyer route
# has been timed from its workbook coordinates and 45 km/h speed.
LAUNCH_RULES = {
    "USV TN-0302": {
        "kind": "absolute",
        "time_s": 0.0,
        "description": "T0: launch from enemy destroyer at D0",
    },
    "USV TN-0301": {
        "kind": "absolute",
        "time_s": 87.0,
        "description": "T0 + 87 s: launch from moving enemy destroyer",
    },
    "UAV TN-0201": {
        "kind": "enemy_waypoint",
        "waypoint": "D1",
        "description": "Launch when enemy destroyer reaches D1",
    },
    "Tên lửa YJ-83": {
        "kind": "enemy_waypoint",
        "waypoint": "D2",
        "description": "Launch when enemy destroyer reaches D2",
    },
}


# ---------------------------------------------------------------------------
# Parsing and geometry
# ---------------------------------------------------------------------------


def normalize_text(value: object) -> str:
    """Normalize whitespace while preserving Vietnamese text for display."""
    return " ".join(str(value or "").strip().split())


def search_key(value: object) -> str:
    """Accent-insensitive lowercase key used only for row/reference matching."""
    text = normalize_text(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("Đ", "D").replace("đ", "d")
    return text.lower()


def canonical_entity_name(value: object) -> str:
    """Map workbook labels to stable names used by the launch schedule."""
    text = normalize_text(value)
    key = search_key(text)

    if "tn-0302" in key:
        return "USV TN-0302"
    if "tn-0301" in key:
        return "USV TN-0301"
    if "tn-0201" in key:
        return "UAV TN-0201"
    if "tn-0202" in key:
        return "UAV TN-0202 (trinh sát)"
    if "yj-83" in key:
        return "Tên lửa YJ-83"
    if "khu truc" in key and "dich" in key:
        return "Tàu khu trục địch"
    if key in {"tau ta", "own ship", "own vessel"}:
        return "Tàu ta"
    return text


def parse_coordinate(value: object) -> Optional[Tuple[float, float]]:
    """Parse '109.510793°W / 11.951853°N' as (latitude, longitude)."""
    if not isinstance(value, str):
        return None

    pattern = (
        r"([0-9]+(?:\.[0-9]+)?)\s*°?\s*([EW])"
        r"\s*/\s*"
        r"([0-9]+(?:\.[0-9]+)?)\s*°?\s*([NS])"
    )
    match = re.search(pattern, value.strip(), flags=re.IGNORECASE)
    if not match:
        return None

    lon = float(match.group(1))
    if match.group(2).upper() == "W":
        lon = -lon

    lat = float(match.group(3))
    if match.group(4).upper() == "S":
        lat = -lat

    return lat, lon


def parse_number(value: object, *, default: Optional[float] = None) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        match = re.search(r"[-+]?[0-9]*\.?[0-9]+", value)
        if match:
            return float(match.group())
    if default is not None:
        return default
    raise ValueError(f"Cannot parse numeric value from {value!r}")


def haversine_km(
    p1: Tuple[float, float],
    p2: Tuple[float, float],
) -> float:
    """Great-circle surface distance in kilometres."""
    lat1, lon1 = p1
    lat2, lon2 = p2

    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)

    a = (
        math.sin(dphi / 2.0) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2.0) ** 2
    )
    return 2.0 * EARTH_RADIUS_KM * math.asin(math.sqrt(a))


def distance_3d_km(
    p1: Tuple[float, float, float],
    p2: Tuple[float, float, float],
) -> float:
    horizontal = haversine_km((p1[0], p1[1]), (p2[0], p2[1]))
    vertical = (p2[2] - p1[2]) / 1000.0
    return math.hypot(horizontal, vertical)


def latlon_to_local_km(
    lat: float,
    lon: float,
    origin_lat: float,
    origin_lon: float,
) -> Tuple[float, float]:
    """Equirectangular local coordinates: X east, Y north, in km."""
    lat_rad = math.radians(lat)
    lon_rad = math.radians(lon)
    lat0_rad = math.radians(origin_lat)
    lon0_rad = math.radians(origin_lon)

    x = EARTH_RADIUS_KM * math.cos(lat0_rad) * (lon_rad - lon0_rad)
    y = EARTH_RADIUS_KM * (lat_rad - lat0_rad)
    return x, y


def interpolate_latlon(
    p1: Tuple[float, float],
    p2: Tuple[float, float],
    fraction: float,
) -> Tuple[float, float]:
    """Linear interpolation, adequate for the short individual route legs."""
    fraction = min(1.0, max(0.0, fraction))
    return (
        p1[0] + fraction * (p2[0] - p1[0]),
        p1[1] + fraction * (p2[1] - p1[1]),
    )


def format_elapsed(seconds: float) -> str:
    total = int(round(seconds))
    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours:d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:d}:{secs:02d}"


# ---------------------------------------------------------------------------
# Workbook loading and route construction
# ---------------------------------------------------------------------------


def find_scenario_header_row(ws) -> int:
    for row in range(1, ws.max_row + 1):
        value = search_key(ws.cell(row, 1).value)
        if "don vi" in value and "ten" in value:
            return row
    raise ValueError("Could not locate the detailed scenario-table header row.")


def raw_waypoint_values(ws, row: int, header_row: int) -> Dict[str, object]:
    # The waypoint columns immediately follow the first four descriptive fields.
    return {
        label: ws.cell(row, 5 + index).value
        for index, label in enumerate(WAYPOINT_LABELS)
    }


def resolve_reference(
    raw: object,
    *,
    own_coord: Tuple[float, float],
    enemy_waypoints: Dict[str, Tuple[float, float]],
) -> Optional[Tuple[float, float]]:
    coord = parse_coordinate(raw)
    if coord is not None:
        return coord

    key = search_key(raw)
    if not key or key == "-":
        return None
    if key in {"tau ta", "own ship", "own vessel"}:
        return own_coord
    if "khu truc" in key:
        if "d0" in key:
            return enemy_waypoints["D0"]
        if "d1" in key:
            return enemy_waypoints["D1"]
        if "d2" in key:
            return enemy_waypoints["D2"]
        if "d3" in key:
            return enemy_waypoints["D3"]
        if "d4" in key:
            return enemy_waypoints["D4"]
    return None


def time_route_points(
    *,
    name: str,
    role: str,
    speed_kmh: float,
    launch_time_s: float,
    launch_rule: str,
    raw_points: Iterable[Tuple[str, float, float, float]],
    own_coord: Tuple[float, float],
) -> dict:
    points: List[dict] = []
    local_elapsed_s = 0.0
    cumulative_distance_km = 0.0
    raw_points = list(raw_points)

    for index, (label, lat, lon, altitude_m) in enumerate(raw_points):
        if index == 0:
            segment_distance_km = 0.0
            segment_time_s = 0.0
        else:
            previous = raw_points[index - 1]
            segment_distance_km = distance_3d_km(
                (previous[1], previous[2], previous[3]),
                (lat, lon, altitude_m),
            )
            segment_time_s = segment_distance_km / speed_kmh * 3600.0
            local_elapsed_s += segment_time_s
            cumulative_distance_km += segment_distance_km

        x_km, y_km = latlon_to_local_km(
            lat,
            lon,
            own_coord[0],
            own_coord[1],
        )

        points.append(
            {
                "label": label,
                "lat": lat,
                "lon": lon,
                "altitude_m": altitude_m,
                "x_km": x_km,
                "y_km": y_km,
                "segment_distance_km": segment_distance_km,
                "segment_time_s": segment_time_s,
                "cumulative_distance_km": cumulative_distance_km,
                "local_elapsed_s": local_elapsed_s,
                "global_time_s": launch_time_s + local_elapsed_s,
            }
        )

    return {
        "name": name,
        "role": role,
        "speed_kmh": speed_kmh,
        "launch_time_s": launch_time_s,
        "launch_rule": launch_rule,
        "points": points,
        "travel_time_s": local_elapsed_s,
        "total_distance_km": cumulative_distance_km,
        "arrival_time_s": launch_time_s + local_elapsed_s,
    }


def position_on_route_at(route: dict, global_time_s: float) -> Tuple[float, float]:
    """Interpolate the moving enemy ship's position at an absolute scenario time."""
    points = route["points"]
    if global_time_s <= points[0]["global_time_s"]:
        return points[0]["lat"], points[0]["lon"]

    for left, right in zip(points, points[1:]):
        t0 = left["global_time_s"]
        t1 = right["global_time_s"]
        if t0 <= global_time_s <= t1:
            span = max(t1 - t0, 1e-12)
            fraction = (global_time_s - t0) / span
            return interpolate_latlon(
                (left["lat"], left["lon"]),
                (right["lat"], right["lon"]),
                fraction,
            )

    return points[-1]["lat"], points[-1]["lon"]


def load_routes(
    workbook_path: Path,
    *,
    include_recon: bool = False,
) -> Tuple[Tuple[float, float], dict, List[dict], List[dict]]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Sheet1"]
    header_row = find_scenario_header_row(ws)

    rows_by_name: Dict[str, int] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        name = canonical_entity_name(ws.cell(row, 1).value)
        if name:
            rows_by_name[name] = row

    required_names = {
        "Tàu ta",
        "Tàu khu trục địch",
        *LAUNCH_RULES.keys(),
    }
    missing = sorted(required_names - rows_by_name.keys())
    if missing:
        raise ValueError(f"Required workbook rows are missing: {missing}")

    own_row = rows_by_name["Tàu ta"]
    own_coord = parse_coordinate(ws.cell(own_row, 5).value)
    if own_coord is None:
        raise ValueError("Could not parse the own-ship coordinate.")

    # Build and time the moving enemy destroyer first.
    enemy_row = rows_by_name["Tàu khu trục địch"]
    enemy_speed = parse_number(ws.cell(enemy_row, 2).value)
    enemy_altitude = parse_number(ws.cell(enemy_row, 3).value, default=0.0)
    enemy_raw_values = raw_waypoint_values(ws, enemy_row, header_row)

    enemy_waypoints: Dict[str, Tuple[float, float]] = {}
    for label in WAYPOINT_LABELS:
        raw = enemy_raw_values[label]
        coord = parse_coordinate(raw)
        if coord is None and search_key(raw) in {"tau ta", "own ship", "own vessel"}:
            coord = own_coord
        if coord is None:
            raise ValueError(f"Could not resolve enemy destroyer {label}: {raw!r}")
        enemy_waypoints[label] = coord

    enemy_route = time_route_points(
        name="Tàu khu trục địch",
        role="enemy_ship",
        speed_kmh=enemy_speed,
        launch_time_s=0.0,
        launch_rule="T0: enemy destroyer starts moving from D0",
        raw_points=[
            (label, enemy_waypoints[label][0], enemy_waypoints[label][1], enemy_altitude)
            for label in WAYPOINT_LABELS
        ],
        own_coord=own_coord,
    )

    enemy_time_by_waypoint = {
        point["label"]: point["global_time_s"]
        for point in enemy_route["points"]
    }

    attack_routes: List[dict] = []
    launch_events: List[dict] = []

    names_to_build = list(LAUNCH_RULES.keys())
    if include_recon and "UAV TN-0202 (trinh sát)" in rows_by_name:
        names_to_build.append("UAV TN-0202 (trinh sát)")

    for name in names_to_build:
        row = rows_by_name[name]
        speed_kmh = parse_number(ws.cell(row, 2).value)
        altitude_m = parse_number(ws.cell(row, 3).value, default=0.0)

        if name in LAUNCH_RULES:
            rule = LAUNCH_RULES[name]
            if rule["kind"] == "absolute":
                launch_time_s = float(rule["time_s"])
            elif rule["kind"] == "enemy_waypoint":
                launch_time_s = enemy_time_by_waypoint[rule["waypoint"]]
            else:
                raise ValueError(f"Unknown launch-rule kind for {name}: {rule}")
            launch_rule = str(rule["description"])
            launch_coord = position_on_route_at(enemy_route, launch_time_s)
        else:
            # The reconnaissance UAV has no launch time in the supplied schedule.
            # It is included only when explicitly requested and starts at T0.
            launch_time_s = 0.0
            launch_rule = "Unscheduled reconnaissance asset; plotted from T0"
            launch_coord = enemy_waypoints["D0"]

        raw_values = raw_waypoint_values(ws, row, header_row)
        route_points: List[Tuple[str, float, float, float]] = [
            ("D0", launch_coord[0], launch_coord[1], altitude_m)
        ]

        # D0 is replaced by the moving launch platform's actual position. Keep
        # the asset's D1-D4 route geometry from the workbook.
        for label in WAYPOINT_LABELS[1:]:
            raw = raw_values[label]
            coord = resolve_reference(
                raw,
                own_coord=own_coord,
                enemy_waypoints=enemy_waypoints,
            )
            if coord is None:
                continue

            # When an airborne target reaches the own ship, end at ship altitude.
            point_altitude = (
                0.0
                if search_key(raw) in {"tau ta", "own ship", "own vessel"}
                else altitude_m
            )
            route_points.append((label, coord[0], coord[1], point_altitude))

        route = time_route_points(
            name=name,
            role="attack_asset" if name in LAUNCH_RULES else "recon_asset",
            speed_kmh=speed_kmh,
            launch_time_s=launch_time_s,
            launch_rule=launch_rule,
            raw_points=route_points,
            own_coord=own_coord,
        )
        attack_routes.append(route)

        launch_events.append(
            {
                "name": name,
                "time_s": launch_time_s,
                "lat": launch_coord[0],
                "lon": launch_coord[1],
                "x_km": route["points"][0]["x_km"],
                "y_km": route["points"][0]["y_km"],
                "rule": launch_rule,
            }
        )

    return own_coord, enemy_route, attack_routes, launch_events


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------


def write_timeline_csv(
    routes: List[dict],
    output_path: Path,
) -> None:
    fields = [
        "entity",
        "role",
        "speed_kmh",
        "launch_rule",
        "launch_time_s",
        "launch_time_hms",
        "waypoint",
        "latitude_deg",
        "longitude_deg",
        "altitude_m",
        "x_east_km",
        "y_north_km",
        "segment_distance_km",
        "segment_time_s",
        "cumulative_distance_km",
        "local_elapsed_s",
        "local_elapsed_hms",
        "global_time_s",
        "global_time_hms",
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()

        for route in routes:
            for point in route["points"]:
                writer.writerow(
                    {
                        "entity": route["name"],
                        "role": route["role"],
                        "speed_kmh": round(route["speed_kmh"], 3),
                        "launch_rule": route["launch_rule"],
                        "launch_time_s": round(route["launch_time_s"], 3),
                        "launch_time_hms": format_elapsed(route["launch_time_s"]),
                        "waypoint": point["label"],
                        "latitude_deg": round(point["lat"], 8),
                        "longitude_deg": round(point["lon"], 8),
                        "altitude_m": round(point["altitude_m"], 3),
                        "x_east_km": round(point["x_km"], 4),
                        "y_north_km": round(point["y_km"], 4),
                        "segment_distance_km": round(point["segment_distance_km"], 4),
                        "segment_time_s": round(point["segment_time_s"], 3),
                        "cumulative_distance_km": round(point["cumulative_distance_km"], 4),
                        "local_elapsed_s": round(point["local_elapsed_s"], 3),
                        "local_elapsed_hms": format_elapsed(point["local_elapsed_s"]),
                        "global_time_s": round(point["global_time_s"], 3),
                        "global_time_hms": format_elapsed(point["global_time_s"]),
                    }
                )


def write_timing_summary_csv(
    attack_routes: List[dict],
    output_path: Path,
) -> None:
    fields = [
        "target",
        "launch_rule",
        "launch_time_s",
        "launch_time_hms",
        "calculated_distance_km",
        "estimated_distance_km",
        "distance_difference_km",
        "calculated_travel_time_s",
        "calculated_travel_time_hms",
        "estimated_travel_time_s",
        "estimated_travel_time_hms",
        "travel_time_difference_s",
        "global_arrival_time_s",
        "global_arrival_time_hms",
    ]

    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()

        for route in attack_routes:
            estimate = ESTIMATED_TRAVEL.get(route["name"], {})
            estimated_distance = estimate.get("distance_km")
            estimated_time = estimate.get("travel_time_s")

            writer.writerow(
                {
                    "target": route["name"],
                    "launch_rule": route["launch_rule"],
                    "launch_time_s": round(route["launch_time_s"], 3),
                    "launch_time_hms": format_elapsed(route["launch_time_s"]),
                    "calculated_distance_km": round(route["total_distance_km"], 4),
                    "estimated_distance_km": estimated_distance,
                    "distance_difference_km": (
                        round(route["total_distance_km"] - estimated_distance, 4)
                        if estimated_distance is not None
                        else ""
                    ),
                    "calculated_travel_time_s": round(route["travel_time_s"], 3),
                    "calculated_travel_time_hms": format_elapsed(route["travel_time_s"]),
                    "estimated_travel_time_s": estimated_time,
                    "estimated_travel_time_hms": (
                        format_elapsed(estimated_time) if estimated_time is not None else ""
                    ),
                    "travel_time_difference_s": (
                        round(route["travel_time_s"] - estimated_time, 3)
                        if estimated_time is not None
                        else ""
                    ),
                    "global_arrival_time_s": round(route["arrival_time_s"], 3),
                    "global_arrival_time_hms": format_elapsed(route["arrival_time_s"]),
                }
            )


# ---------------------------------------------------------------------------
# Plots
# ---------------------------------------------------------------------------


def plot_routes(
    own_coord: Tuple[float, float],
    enemy_route: dict,
    attack_routes: List[dict],
    launch_events: List[dict],
    output_path: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(14, 9))

    ax.scatter([0.0], [0.0], marker="*", s=260, label="Own ship", zorder=12)

    # Moving enemy destroyer route.
    ex = [point["x_km"] for point in enemy_route["points"]]
    ey = [point["y_km"] for point in enemy_route["points"]]
    enemy_line = ax.plot(
        ex,
        ey,
        marker="s",
        linewidth=2.4,
        markersize=6,
        color="black",
        label=(
            f'{enemy_route["name"]} '
            f'({enemy_route["speed_kmh"]:g} km/h)'
        ),
        zorder=7,
    )[0]

    enemy_offsets = {
        "D0": (8, 8),
        "D1": (8, 8),
        "D2": (8, 14),
        "D3": (8, 8),
    }
    for point in enemy_route["points"]:
        if point["label"] == "D4":
            continue
        ax.annotate(
            f'{point["label"]}\nT={format_elapsed(point["global_time_s"])}',
            (point["x_km"], point["y_km"]),
            xytext=enemy_offsets.get(point["label"], (6, 6)),
            textcoords="offset points",
            fontsize=7.5,
            color=enemy_line.get_color(),
            fontweight="bold",
        )

    # Attack-asset routes.
    route_color: Dict[str, object] = {}
    for route in attack_routes:
        xs = [point["x_km"] for point in route["points"]]
        ys = [point["y_km"] for point in route["points"]]

        line = ax.plot(
            xs,
            ys,
            marker="o",
            linewidth=1.9,
            markersize=5,
            label=(
                f'{route["name"]} ({route["speed_kmh"]:g} km/h; '
                f'launch {format_elapsed(route["launch_time_s"])})'
            ),
            zorder=5,
        )[0]
        route_color[route["name"]] = line.get_color()

        for point in route["points"][:-1]:
            ax.annotate(
                f'{point["label"]}\nT={format_elapsed(point["global_time_s"])}',
                (point["x_km"], point["y_km"]),
                xytext=(5, 5),
                textcoords="offset points",
                fontsize=7,
                color=line.get_color(),
            )

    # Explicit launch markers, including the moving launch position at +87 s.
    launch_offsets = {
        "USV TN-0302": (10, 16),
        "USV TN-0301": (10, -30),
        "UAV TN-0201": (10, -30),
        "Tên lửa YJ-83": (10, -34),
    }
    for event in launch_events:
        color = route_color.get(event["name"], "tab:red")
        ax.scatter(
            [event["x_km"]],
            [event["y_km"]],
            marker="*",
            s=150,
            c=[color],
            edgecolors="black",
            linewidths=0.7,
            zorder=11,
        )
        ax.annotate(
            f'{event["name"]} launch\nT={format_elapsed(event["time_s"])}',
            (event["x_km"], event["y_km"]),
            xytext=launch_offsets.get(event["name"], (8, -20)),
            textcoords="offset points",
            fontsize=7.5,
            color=color,
            bbox=dict(facecolor="white", alpha=0.78, edgecolor=color),
        )

    # Separate arrival callouts around the common own-ship endpoint.
    arrival_offsets = {
        "USV TN-0302": (12, -18),
        "USV TN-0301": (12, 0),
        "UAV TN-0201": (12, 18),
        "Tên lửa YJ-83": (12, 36),
    }
    for route in attack_routes:
        final = route["points"][-1]
        color = route_color[route["name"]]
        ax.annotate(
            f'{route["name"]} arrival {format_elapsed(route["arrival_time_s"])}',
            (final["x_km"], final["y_km"]),
            xytext=arrival_offsets.get(route["name"], (12, 0)),
            textcoords="offset points",
            fontsize=7.2,
            color=color,
            ha="left",
            va="center",
        )

    ax.set_title("Moving enemy destroyer, launch points, and target waypoint routes")
    ax.set_xlabel("East displacement from own ship (km)")
    ax.set_ylabel("North displacement from own ship (km)")
    ax.axhline(0.0, linewidth=0.8, color="gray")
    ax.axvline(0.0, linewidth=0.8, color="gray")
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.axis("equal")
    ax.legend(loc="best", fontsize=8)
    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def plot_distance_to_ship(
    routes: List[dict],
    output_path: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(13, 7))

    for route in routes:
        times_min = [point["global_time_s"] / 60.0 for point in route["points"]]
        distances = [
            math.sqrt(
                point["x_km"] ** 2
                + point["y_km"] ** 2
                + (point["altitude_m"] / 1000.0) ** 2
            )
            for point in route["points"]
        ]

        line = ax.plot(
            times_min,
            distances,
            marker="o",
            linewidth=1.9,
            label=route["name"],
        )[0]

        for point, time_min, distance in zip(route["points"], times_min, distances):
            ax.annotate(
                point["label"],
                (time_min, distance),
                xytext=(4, 4),
                textcoords="offset points",
                fontsize=7,
                color=line.get_color(),
            )

    ax.set_title("Distance to own ship on the global operational timeline")
    ax.set_xlabel("Global elapsed time from T0 (minutes)")
    ax.set_ylabel("Distance to own ship (km)")
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.legend(loc="best", fontsize=8)
    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def plot_launch_arrival_timeline(
    enemy_route: dict,
    attack_routes: List[dict],
    output_path: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(13, 6.5))

    ordered = sorted(attack_routes, key=lambda route: route["launch_time_s"])
    y_positions = list(range(len(ordered)))

    for y, route in zip(y_positions, ordered):
        launch_min = route["launch_time_s"] / 60.0
        arrival_min = route["arrival_time_s"] / 60.0
        ax.hlines(y, launch_min, arrival_min, linewidth=5, alpha=0.75)
        ax.scatter([launch_min], [y], marker="|", s=180, linewidths=2.5)
        ax.scatter([arrival_min], [y], marker="o", s=55)
        ax.annotate(
            f'launch {format_elapsed(route["launch_time_s"])}',
            (launch_min, y),
            xytext=(0, -16 if route["name"] == "Tên lửa YJ-83" else 10),
            textcoords="offset points",
            fontsize=7,
            ha="left",
            va="center",
        )
        ax.annotate(
            f'arrive {format_elapsed(route["arrival_time_s"])}',
            (arrival_min, y),
            xytext=(0, 12),
            textcoords="offset points",
            fontsize=7,
            ha="right",
            va="center",
        )

    enemy_times = {
        point["label"]: point["global_time_s"] / 60.0
        for point in enemy_route["points"]
    }
    for label in ("D1", "D2"):
        x = enemy_times[label]
        ax.axvline(x, linestyle="--", linewidth=1.2, alpha=0.7)
        ax.text(
            x,
            len(ordered) - 0.15,
            f'Enemy {label}\n{format_elapsed(x * 60.0)}',
            rotation=90,
            va="top",
            ha="right",
            fontsize=7.5,
        )

    ax.set_yticks(y_positions)
    ax.set_yticklabels([route["name"] for route in ordered])
    ax.set_xlabel("Global elapsed time from T0 (minutes)")
    ax.set_title("Attack launch and calculated arrival timeline")
    ax.grid(axis="x", linestyle="--", alpha=0.35)
    ax.set_ylim(-0.6, len(ordered) - 0.05)
    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Plot the moving enemy destroyer and scheduled target launches "
            "using the updated scenario workbook."
        )
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default="scenario.xlsx",
        help="Path to the updated scenario workbook.",
    )
    parser.add_argument(
        "--outdir",
        default="waypoint_output",
        help="Directory for generated CSV and PNG files.",
    )
    parser.add_argument(
        "--include-recon",
        action="store_true",
        help=(
            "Also plot UAV TN-0202. It has no launch time in the supplied "
            "schedule and is therefore plotted from T0."
        ),
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    output_dir = Path(args.outdir)
    output_dir.mkdir(parents=True, exist_ok=True)

    own_coord, enemy_route, attack_routes, launch_events = load_routes(
        workbook_path,
        include_recon=args.include_recon,
    )

    all_routes = [enemy_route, *attack_routes]

    timeline_csv = output_dir / "waypoint_timeline.csv"
    summary_csv = output_dir / "timing_summary.csv"
    route_plot = output_dir / "waypoint_routes.png"
    distance_plot = output_dir / "distance_to_own_ship.png"
    timeline_plot = output_dir / "launch_arrival_timeline.png"

    write_timeline_csv(all_routes, timeline_csv)
    write_timing_summary_csv(attack_routes, summary_csv)
    plot_routes(own_coord, enemy_route, attack_routes, launch_events, route_plot)
    plot_distance_to_ship(all_routes, distance_plot)
    plot_launch_arrival_timeline(enemy_route, attack_routes, timeline_plot)

    print(f"Created: {timeline_csv}")
    print(f"Created: {summary_csv}")
    print(f"Created: {route_plot}")
    print(f"Created: {distance_plot}")
    print(f"Created: {timeline_plot}")

    print("\nEnemy destroyer waypoint times:")
    for point in enemy_route["points"]:
        print(
            f'  {point["label"]}: T={format_elapsed(point["global_time_s"])} '
            f'({point["global_time_s"]:.3f} s)'
        )

    print("\nAttack timing summary:")
    for route in sorted(attack_routes, key=lambda item: item["launch_time_s"]):
        estimate = ESTIMATED_TRAVEL.get(route["name"])
        estimate_text = ""
        if estimate:
            delta = route["travel_time_s"] - estimate["travel_time_s"]
            estimate_text = (
                f'; supplied estimate={format_elapsed(estimate["travel_time_s"])}'
                f'; delta={delta:+.1f} s'
            )

        print(
            f'  {route["name"]}: launch={format_elapsed(route["launch_time_s"])}; '
            f'distance={route["total_distance_km"]:.2f} km; '
            f'travel={format_elapsed(route["travel_time_s"])}; '
            f'arrival={format_elapsed(route["arrival_time_s"])}'
            f'{estimate_text}'
        )

    print(
        "\nNote: USV TN-0301 starts from the enemy destroyer's interpolated "
        "position at T0+87 s, not from the fixed D0 coordinate."
    )


if __name__ == "__main__":
    main()
